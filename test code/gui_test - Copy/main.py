import PySimpleGUI as S_gui
import serial  # Check if needed
import warnings
import serial.tools.list_ports
from datetime import datetime
import time
import json
import os

# Great site: https://pysimplegui.readthedocs.io/en/latest/call%20reference/#button-element

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.cell import get_column_letter
import csv

version = '1.0.0'

showThemeSelect = False
theme = default_theme = flow = existing_file_path = existing_file = convert_to_excel = root_folder = None


def estimate_width(value):
    _value = str(value)
    return len(_value)


def convert_csv_to_excel(csv_path):
    skip = 0
    excel_name = csv_path[:-4] + '.xlsx'
    if os.path.isfile(excel_name):
        wb = load_workbook(excel_name)
        skip = wb.active.max_row
    else:
        wb = Workbook()
    ws = wb.active
    counter = 0

    with open(csv_path, 'r') as f:
        for _row in csv.reader(f):
            if counter == skip:
                _list = []
                for item in _row:
                    try:
                        _list.append(int(item))
                    except:
                        try:
                            _list.append(float(item))
                        except:
                            _list.append(str(item))
                ws.append(_list)
            else:
                counter = counter + 1
    for _row in range(skip + 1, ws.max_row + 1):
        for _col in range(1, ws.max_column + 1):
            active_cell = ws.cell(_row, _col)
            active_cell.alignment = Alignment(horizontal='right')
    if skip == 0:
        col = 1
        for _cell in ws["1:1"]:
            _cell.font = Font(bold=True)
            column_size = max(estimate_width(ws.cell(row=1, column=col).value),
                              estimate_width(ws.cell(row=2, column=col).value)) + 3
            ws.column_dimensions[get_column_letter(col)].width = column_size
            col = col + 1
    wb.save(excel_name)


# Reading settings
def read_settings():
    global showThemeSelect, theme, default_theme, flow, existing_file_path, existing_file, convert_to_excel, root_folder
    with open('settings.json', 'r') as read_settings_file:
        settings_data = json.load(read_settings_file)

    showThemeSelect = settings_data['showThemeSelect']
    S_gui.theme(settings_data['defaultTheme'])
    default_theme = theme = settings_data['defaultTheme']
    flow = settings_data['flow']
    existing_file = settings_data['WriteToAnExistingFile']
    existing_file_path = settings_data['WriteToAnExistingFilePath']
    convert_to_excel = settings_data['ConvertToExcel']
    root_folder = settings_data['RootFolder']


read_settings()


def update_settings(_flow, _pop_values):
    global success, showThemeSelect, default_theme, flow, existing_file, convert_to_excel, root_folder
    with open('settings.json', 'r') as read_file:
        data = json.load(read_file)
    data['flow'] = _flow
    if _pop_values['-DEFAULT-']:
        data['defaultTheme'] = theme
        default_theme = theme
    showThemeSelect = _pop_values['-SHOW-']
    data['showThemeSelect'] = showThemeSelect
    existing_file = _pop_values['-FILE-']
    data['WriteToAnExistingFile'] = existing_file
    if _pop_values['-FILE-']:
        data['WriteToAnExistingFilePath'] = existing_file_path
    data['RootFolder'] = root_folder
    convert_to_excel = _pop_values['-EXCEL-']
    data['ConvertToExcel'] = convert_to_excel
    with open('settings.json', 'w') as _write_settings_file:
        json.dump(data, _write_settings_file, indent=4)


# Theme selector
if showThemeSelect:
    S_gui.popup_quick_message('Hang on for a moment, this will take a bit to create...', background_color='red',
                              text_color='white', auto_close=True, non_blocking=True)
    temp_layout = []
    names = S_gui.list_of_look_and_feel_values()
    names.sort()
    row = []
    for count, theme in enumerate(names):
        S_gui.change_look_and_feel(theme)
        if not count % 6:
            temp_layout += [row]
            row = []
        row += [
            S_gui.Frame(theme,
                        [[S_gui.Text('Preview for this theme:'), S_gui.InputText('Input data here', size=(10, 1))],
                         [S_gui.Button('Ok', k=theme), S_gui.Button('Cancel', disabled=True)]])]
    if row:
        temp_layout += [row]

    select_win = S_gui.Window('Select a theme!',
                              [[S_gui.Column(temp_layout, scrollable=True, vertical_scroll_only=True)]],
                              element_justification='Center', finalize=True)
    select_win.maximize()
    selected, num2 = select_win.read()
    S_gui.theme(selected)
    theme = selected
    select_win.close()

# Connections variables
file = lines = 0
list_v = [-1.0, -1.0, -1.0, -1.0, -1.0]
input_s = input_e = input_f = ""
add_to_file = False
writing_entries = False
recording_file = None
read = False
good = False
scenario_file_name = ''
cancel_scenario = False
run_through_scenario = False
first_run = True
index = 0
start = 0
end_time = 0
recording_timer = 0
csv_file_name = "TempName.csv"

# More variables
delay_timer = faucet1_timer = faucet2_timer = 0
faucet1_timer_active = False
faucet2_timer_active = False
faucet1_cond_active = False
faucet1_parameter = ""
faucet1_mode = ""
faucet1_stop_value = 0
faucet2_cond_active = False
faucet2_parameter = ""
faucet2_mode = ""
faucet2_stop_value = 0
recording_rate = 1  # Rate of recording, in seconds

cond_dic = {
    'light': 2,
    'temp': 0,
    'moisture': 3,
    'humidity': 1
}

mode_dic = {
    '>': True,
    '<': False
}

options_dic = {
    'פעם ב-10 דקות': 10,
    'פעם ב-5 דקות': 5,
    'פעם בדקה': 1
}


def start_recording_file():
    global writing_entries, csv_file_name, recording_file
    writing_entries = True
    csv_file_name = root_folder + "\\Recording_file_" + str(int(time.time())) + ".csv"
    if existing_file:
        recording_file = open(existing_file_path, "a")
        csv_file_name = existing_file_path
    else:
        recording_file = open(csv_file_name, "w")
        recording_file.write(
            'זמן' + ',' + 'טמפרטורה' + ',' + 'לחות - אויר' + ',' + 'אור' + ',' + 'לחות - אדמה' + ',' + 'ph' + '\n')
    window['-RECORD-'].update(image_filename='cancel.png')


def stop_recording_file():
    global writing_entries, recording_file
    writing_entries = False
    recording_file.close()
    window['-RECORD-'].update(image_filename='record-image.png', image_subsample=3)


def faucet1_set_cond(parameter, mode, value):
    global faucet1_cond_active, faucet1_parameter, faucet1_mode, faucet1_stop_value
    faucet1_parameter = parameter
    faucet1_mode = mode
    faucet1_stop_value = int(value)
    faucet1_cond_active = True
    open_faucet1()


def faucet2_set_cond(parameter, mode, value):
    global faucet2_cond_active, faucet2_parameter, faucet2_mode, faucet2_stop_value
    faucet2_parameter = parameter
    faucet2_mode = mode
    faucet2_stop_value = int(value)
    faucet2_cond_active = True
    open_faucet2()


faucet_dic_cond = {
    '1': faucet1_set_cond,
    '2': faucet2_set_cond,
}


def print_line(input_string):
    window['-OUTPUT_T-'].update(input_string)
    apply_scenario()


def delay_line(input_string):
    global delay_timer
    delay_timer = int(time.time()) + int(input_string)


def faucet_line(input_string):
    global faucet1_timer_active, faucet2_timer_active, faucet1_timer, faucet2_timer, flow
    variables = input_string.split(' ')
    if variables[1] == 't':
        if variables[0] == '1':
            faucet1_timer = int(time.time()) + int(variables[2])
            open_faucet1()
            faucet1_timer_active = True
        else:
            faucet2_timer = int(time.time()) + int(variables[2])
            open_faucet2()
            faucet2_timer_active = True
    elif variables[1] == 'v':
        faucet_dic_cond[variables[0]](variables[2], variables[3], variables[4])
    elif variables[1] == 'f':
        if variables[0] == '1':
            faucet1_timer = int(time.time()) + (int(variables[2]) / ((flow * 1000) / 3600))
            open_faucet1()
            faucet1_timer_active = True
        else:
            faucet2_timer = int(time.time()) + int(variables[2] * 1000 * 3600)
            open_faucet2()
            faucet2_timer_active = True


def color_line(input_string):
    global red, green, blue
    split = input_string.split()
    red = split[0]
    green = split[1]
    blue = split[2]
    window['-BLUE-'].update(blue)
    window['-RED-'].update(red)
    window['-GREEN-'].update(green)
    ser.write(bytes(('c ' + str(red) + ' ' + str(green) + ' ' + str(blue) + '\n').encode('utf8', 'strict')))


scenario_dic = {
    '@': print_line,
    'o': faucet_line,
    's': color_line,
    'd': delay_line,
}

# Define the window's contents
red = green = blue = 0
prev_value = '#000000'
down = True

# Faucet control
f_water = f_fertilizer = True


def apply_scenario():
    global first_run, index, start, delay_timer, recording_rate, end_time, file, lines
    if first_run:
        file = open(scenario_file_name, 'r')
        lines = file.readlines()
        start = finish = 0
        first_run = False
        for loop_i in range(0, len(lines)):
            if lines[loop_i][0] == '{':
                start = loop_i + 1
            elif lines[loop_i][0] == '}':
                finish = loop_i - 1
                break
        recording_rate = int(lines[start])
        index = start = finish + 2
        end_time = time.time() + 86400
    if index < len(lines):
        index = index + 1
        delay_timer = time.time() + 1
        scenario_dic[lines[index - 1][0]](lines[index - 1][2:])
    if end_time <= time.time():
        index = start


def open_faucet1():
    global f_water
    f_water = False
    window['-WATER-'].update(image_filename='faucet-image-on.png', image_size=(150, 150),
                             image_subsample=3)
    ser.write(bytes(b'3\n'))


def open_faucet2():
    global f_fertilizer
    f_fertilizer = False
    window['-FERTILIZER-'].update(image_filename='fertilizer-png-on.png', image_size=(150, 150),
                                  image_subsample=3)
    ser.write(bytes(b'5\n'))


def close_faucet1():
    global f_water
    f_water = True
    window['-WATER-'].update(image_filename='faucet-image-off.png', image_size=(150, 150),
                             image_subsample=3)
    ser.write(bytes(b'4\n'))


def close_faucet2():
    global f_fertilizer
    f_fertilizer = True
    window['-FERTILIZER-'].update(image_filename='fertilizer-png-off.png', image_size=(150, 150),
                                  image_subsample=3)
    ser.write(bytes(b'6\n'))


def disable_inputs():
    window['-FERTILIZER-'].update(disabled=True)
    window['-WATER-'].update(disabled=True)
    window['-RECORD-'].update(disabled=True)
    window['-OPTIONS_REC-'].update(disabled=True)
    window['-COLOR-'].update(disabled=True)
    window['-UPDATE-'].update(disabled=True)


def enable_inputs():
    window['-FERTILIZER-'].update(disabled=False)
    window['-WATER-'].update(disabled=False)
    window['-RECORD-'].update(disabled=False)
    window['-OPTIONS_REC-'].update(disabled=False)
    window['-COLOR-'].update(disabled=False)
    window['-UPDATE-'].update(disabled=False)


# Establishing a connection with a connected arduino
arduino_ports = [
    p.device
    for p in serial.tools.list_ports.comports()
    if 'Arduino' in p.description  # may need tweaking to match new arduinos
]
while not arduino_ports:
    e, v = S_gui.Window('שגיאה! לא נמצא ארדואינו מחובר', [[S_gui.T('ארדואינו לא מחובר או לא נמצא! נא לבדוק חיבור')],
                                                          [S_gui.Button('בדוק שוב'),
                                                           S_gui.Button('צא', k='--exit--')]],
                        disable_close=True).read(close=True)
    if e == '--exit--':
        exit()
    arduino_ports = [
        p.device
        for p in serial.tools.list_ports.comports()
        if 'Arduino' in p.description  # may need tweaking to match new arduinos
    ]
if len(arduino_ports) > 1:
    warnings.warn('Multiple Arduinos found - using the first')

ser = serial.Serial(arduino_ports[0], 9600, timeout=0, parity=serial.PARITY_NONE,
                    stopbits=serial.STOPBITS_ONE,
                    bytesize=serial.EIGHTBITS,
                    xonxoff=False,
                    rtscts=False,
                    write_timeout=None,
                    dsrdtr=False,
                    inter_byte_timeout=None,
                    exclusive=None)

# The sensor values row
column_names = [
    [S_gui.Text(text='Ph', justification='center', font='david 30 normal', size=(10, 2), pad=(10, 10)),
     S_gui.Text(text='לחות - אדמה', justification='center', size=(10, 2), pad=(14, 10), font='david 30 normal'),
     S_gui.Text(text='אור', justification='center', font='david 30 normal', size=(10, 2), pad=(12, 10)),
     S_gui.Text(text='לחות - אויר', justification='center', font='david 30 normal', size=(10, 2), pad=(12, 10)),
     S_gui.Text(text='טמפרטורה', justification='center', font='david 30 normal', size=(10, 2), pad=(14, 10))],
    # New row, now we show the values
    [S_gui.Frame(title='', relief=S_gui.RELIEF_RAISED, background_color='grey', layout=[
        [S_gui.Text(text='loading...', justification='center', k='-PH-', font='david 30 normal', size=(10, 1))]]),
     S_gui.Frame(title='', relief=S_gui.RELIEF_RAISED, background_color='grey', vertical_alignment='center', layout=[
         [S_gui.Text(text='loading...', justification='center', k='-gHum-', font='david 30 normal', size=(10, 1))]]),
     S_gui.Frame(title='', background_color='grey', relief=S_gui.RELIEF_RAISED, layout=[
         [S_gui.Text(text='loading...', justification='center', k='-light-', font='david 30 normal', size=(10, 1))]]),
     S_gui.Frame(title='', background_color='grey', relief=S_gui.RELIEF_RAISED, layout=[
         [S_gui.Text(text='loading...', justification='center', k='-hum-', font='david 30 normal', size=(10, 1))]]),
     S_gui.Frame(title='', background_color='grey', relief=S_gui.RELIEF_RAISED, layout=[
         [S_gui.Text(text='loading...', justification='center', k='-temp-', font='david 30 normal', size=(10, 1))]])]]

# The RGB controller row
column_light = [[S_gui.Button(button_text='החל', font='david 30 normal', k='-UPDATE-'),
                 S_gui.Input(tooltip='הכניסו ערך בין 0 ל255', justification='center', k='-RED-', size=(10, 1),
                             default_text='0', background_color='red', font='david 30 normal'),
                 S_gui.Input(tooltip='הכניסו ערך בין 0 ל255', justification='center', k='-GREEN-', size=(10, 1),
                             default_text='0', background_color='light green', font='david 30 normal'),
                 S_gui.Input(default_text='0', tooltip='הכניסו ערך בין 0 ל255', justification='center', k='-BLUE-',
                             size=(10, 1), background_color='light blue', font='david 30 normal'),
                 S_gui.Text(' או ', font='david 30 normal'),
                 S_gui.Input(default_text='#000000', visible=False, k='-C-'),
                 S_gui.ColorChooserButton('בחר צבע', image_filename='color.png', image_subsample=5, k='-COLOR-',
                                          target='-C-')]]

column_btnControls = [
    [S_gui.Button(k='-WATER-', image_filename='faucet-image-off.png', image_size=(150, 150),
                  tooltip='לחצו כאן כדי לפתוח או לסגור את ברז 1', image_subsample=3),
     S_gui.Button(k='-FERTILIZER-', image_filename='fertilizer-png-off.png', image_size=(150, 150),
                  tooltip='לחצו כאן כדי לפתוח או לסגור את ברז 2', image_subsample=3)]]

column_recording = [[S_gui.Button(image_filename='record-image.png', image_size=(150, 150), k='-RECORD-',
                                  tooltip='לחצו כאן כדי להתחיל הקלטת נתונים', image_subsample=3),
                     S_gui.OptionMenu(values=["פעם בדקה", "פעם ב-5 דקות", "פעם ב-10 דקות"], s=30, k='-OPTIONS_REC-'),
                     S_gui.Text(text=':תדירות', font='david 30 normal'),
                     S_gui.Button(image_filename='upload-image.png', image_size=(150, 150), k='-UPLOAD-',
                                  tooltip='לחצו כאן כדי לבחור תכנית פעולה')]]

layout = [[S_gui.Column(
    [[S_gui.Text(text='Still loading...', key='-TIME-', justification='left', font='david 16 normal', size=(10, 2),
                 text_color='light grey')]], expand_x=True),
    S_gui.Text(text='', key='-OUTPUT_T-', justification='center', font='david 16 normal',
               text_color='red'),
    S_gui.Button(k='-SETTINGS-', image_filename='settings-png.png', image_size=(50, 50), tooltip='הגדרות',
                 image_subsample=10), S_gui.Button(image_filename='help.png', image_subsample=7, k='-HELP-',
                                                   tooltip='בחירת תיקייה אליה יכתבו קבצי אקסל חדשים')],
    [S_gui.Column([[S_gui.Text(text='בקרת חממה', justification='center', font='david 44 normal', size=(10, 1),
                               text_color='yellow')]], justification='center', vertical_alignment='center',
                  pad=(7, 15))],
    [S_gui.Column([[S_gui.Frame(vertical_alignment='center', title_location=S_gui.TITLE_LOCATION_TOP_RIGHT,
                                font='david 30 normal', title=':ערכי חיישנים',
                                layout=[
                                    [S_gui.Column(column_names, justification='center',
                                                  vertical_alignment='center')]])]], justification='center',
                  pad=(7, 7))],
    [S_gui.Column([[S_gui.Frame(vertical_alignment='center', title_location=S_gui.TITLE_LOCATION_TOP_RIGHT,
                                font='david 30 normal', title=':תאורה', layout=[
            [S_gui.Column(column_light, justification='center', vertical_alignment='center')]])]],
                  justification='center', pad=(7, 7))],
    [S_gui.Column([[S_gui.Frame(vertical_alignment='center', title_location=S_gui.TITLE_LOCATION_TOP_RIGHT,
                                font='david 30 normal', title=':בקרת ברזים', layout=[
            [S_gui.Column(column_btnControls, justification='center', vertical_alignment='center')]]),
                    S_gui.Frame(vertical_alignment='center', title_location=S_gui.TITLE_LOCATION_TOP_RIGHT,
                                font='david 30 normal', title=':אפשרויות הפעלה והקלטה', layout=[
                            [S_gui.Column(column_recording, justification='center', vertical_alignment='center')]])]],
                  justification='center', pad=(7, 7))],
    [S_gui.Button(button_text='יציאה', font='david 30 normal', k='Quit')]]

# Create the window
window = S_gui.Window('Greenhouse GUI', layout, resizable=True, no_titlebar=True, keep_on_top=True).Finalize()
window.Maximize()

# Display and interact with the Window using an Event Loop
while True:
    event, values = window.read(timeout=50)
    # See if user wants to quit or window was closed
    if event == S_gui.WINDOW_CLOSED:
        break
    elif event == '-WATER-':
        if f_water:
            open_faucet1()
        else:
            close_faucet1()
    elif event == '-FERTILIZER-':
        if f_fertilizer:
            open_faucet2()
        else:
            close_faucet2()
    elif event == '-RECORD-':
        if writing_entries:
            stop_recording_file()
            if convert_to_excel:
                convert_csv_to_excel(csv_file_name)
        else:
            try:
                recording_rate = options_dic[values['-OPTIONS_REC-']]
                start_recording_file()
            except:
                S_gui.popup_error('Recording rate not set', keep_on_top=True)
    elif event == '-UPLOAD-':
        if not cancel_scenario:
            scenario_file_name = S_gui.popup_get_file(message='בחר תרחיש', title='Choose scenario', keep_on_top=True)
            if scenario_file_name is not None:
                run_through_scenario = True
                start_recording_file()
                disable_inputs()
                window['-UPLOAD-'].update(image_filename='cancel.png')
                cancel_scenario = True
        else:
            cancel_scenario = False
            run_through_scenario = False
            stop_recording_file()
            first_run = True
            window['-UPLOAD-'].update(image_filename='upload-image.png')
            enable_inputs()

    elif event == '-SETTINGS-':
        success = False
        settings_win = S_gui.Window('הגדרות נוספות', [
            [S_gui.Input(default_text=flow, k='-FLOW-'), S_gui.T(':ספיקה בליטרים לשעה'),
             S_gui.Image(source='help.png', subsample=15, tooltip='הערך קובע את ספיקות הצינור')],
            [S_gui.Checkbox('הגדר ערכת נושא כברירת מחדל', default=True if theme is default_theme else False,
                            k='-DEFAULT-'),
             S_gui.Image(source='help.png', subsample=15, tooltip='אם מסומן ערכת הנושא הנוכחית תהיה ברירת המחדל')],
            [S_gui.Checkbox('הצג בחירת ערכת נושא', default=showThemeSelect, k='-SHOW-'),
             S_gui.Image(source='help.png', subsample=15, tooltip='אם מסומן בהפעלה הבאה יופיע מסך לבחירת ערכת נושא')],
            [S_gui.Checkbox('כתוב לקובץ קיים', default=existing_file, k='-FILE-'),
             S_gui.Image(source='help.png', subsample=15, tooltip='אם מסומן הקלטות ישמרו לקובץ קיים, לפי הקובץ הנבחר')],
            [S_gui.T(text=existing_file_path, k='-FILE_PATH-'),
             S_gui.Button(button_text='בחר קובץ קיים לכתוב אליו', k='-FILE_SELECT-'),
             S_gui.Image(source='help.png', subsample=15,
                         tooltip='בחר אתת הקובץ אליו יכתב המידע אם התוכנית כותבת לקובץ קיים')],
            [S_gui.T(text=root_folder, k='-FOLDER_PATH-'),
             S_gui.Button(button_text='בחר תיקייה אליה ישמרו קבצי האקסל', k='-FOLDER_SELECT-'),
             S_gui.Image(source='help.png', subsample=15, tooltip='בחירת תיקייה אליה יכתבו קבצי אקסל חדשים')],
            [S_gui.Checkbox('המר קובץ לאקסל', default=convert_to_excel, k='-EXCEL-'),
             S_gui.Image(source='help.png', subsample=15, tooltip='אם נבחר יווצר קובץ אקסל בנוסף')],
            [S_gui.Button('החל', s=10, k='-APPLY-'), S_gui.Button('בטל', s=10, k='-CANCEL-')],
            [S_gui.Button('אפס הגדרות', k='-RESET_SETTINGS-')]], element_justification='Right', keep_on_top=True)
        while not success:
            status, pop_values = settings_win.read()
            if status == '-APPLY-':
                try:
                    flow = float(pop_values['-FLOW-'])
                    update_settings(flow, pop_values)
                    success = True
                except:
                    S_gui.popup_error('קלא לא תקין!', keep_on_top=True)
            elif status == '-FILE_SELECT-':
                _existing_file_path = S_gui.popup_get_file(message='בחר קובץ', title='Choose file', keep_on_top=True)
                while _existing_file_path is not None and _existing_file_path[-4:] != '.csv':
                    S_gui.popup_error('בלבד' + ' csv ' + 'נבחר קובץ לא מתאים! יש לבחור קבצי', keep_on_top=True)
                    _existing_file_path = S_gui.popup_get_file(message='בחר קובץ', title='Choose file',
                                                               keep_on_top=True)
                if _existing_file_path is not None:
                    existing_file_path = _existing_file_path
                settings_win['-FILE_PATH-'].update(existing_file_path)
            elif status == '-FOLDER_SELECT-':
                _folder_path = S_gui.popup_get_folder(message='בחר תיקייה', title='Choose folder', keep_on_top=True)
                while _folder_path == "":
                    S_gui.popup_error('נא לבחור תיקייה', keep_on_top=True)
                    _folder_path = S_gui.popup_get_folder(message='בחר תיקייה', title='Choose folder', keep_on_top=True)
                if _folder_path is not None:
                    root_folder = _folder_path
                settings_win['-FOLDER_PATH-'].update(root_folder)

            elif status == '-RESET_SETTINGS-':
                with open('default_settings.json', 'r') as read_file:
                    tempData = json.load(read_file)
                with open('settings.json', 'w') as write_settings_file:
                    json.dump(tempData, write_settings_file, indent=4)
                read_settings()
                success = True
            else:
                success = True
        settings_win.close()

    elif event == '-UPDATE-':
        if prev_value != values['-C-']:
            prev_value = values['-C-']
            tempString = ""
            tempString += prev_value[1]
            tempString += prev_value[2]
            red = int(tempString, 16)
            tempString = ""
            tempString += prev_value[3]
            tempString += prev_value[4]
            green = int(tempString, 16)
            tempString = ""
            tempString += prev_value[5]
            tempString += prev_value[6]
            blue = int(tempString, 16)
            window['-BLUE-'].update(blue)
            window['-RED-'].update(red)
            window['-GREEN-'].update(green)
        else:
            red = values['-RED-']
            blue = values['-BLUE-']
            green = values['-GREEN-']
        ser.write(bytes(('c ' + str(red) + ' ' + str(green) + ' ' + str(blue) + '\n').encode('utf8', 'strict')))

    elif event == 'Quit':
        e_status = S_gui.popup_yes_no("האם אתם בטוחים שאתם רוצים לצאת? אם תצאו הקלטת הנתונים תיפסק",
                                      title='Exit screen', keep_on_top=True)
        if e_status == 'Yes':
            break
    elif event == '-HELP-':
        finished = False
        help_window = S_gui.Window(title='עזרה', layout=[[S_gui.T('זה חלון העזרה, הוא יעזור לכם להתמצא בתוכנה')],
                                                         [S_gui.Button('עזרה עם הקלטה (שמירה לאקסל)', k='-R_help-')],
                                                         [S_gui.Button('הסבר נתונים וערכים', k='-V_help-')],
                                                         [S_gui.Button('הסבר על תרחישים', k='-S_help-')],
                                                         [S_gui.Button('הסבר על השליטה בחממה', k='-C_help-')],
                                                         [S_gui.T(text='V' + version, justification='right')],
                                                         [S_gui.Button('יציאה', k='-Exit-')]],
                                   element_justification='Right', keep_on_top=True)
        while not finished:
            selection, bs_values = help_window.read()
            if selection == '-Exit-':
                finished = True
            elif selection == '-R_help-':
                S_gui.Window(title='עזרה עם הקלטות',
                             layout=[[S_gui.T('?כיצד מקליטים נתונים')],
                                     [S_gui.T(
                                         ':כשרוצים להקליט את הנתונים לקובץ אקסל אפשר להשתמש באחת מהשיטות הבאות')],
                                     [S_gui.T(
                                         'אפשר להתחיל הקלטה מהמסך הראשי על ידי בחירת תדירות הקלטה           ')],
                                     [S_gui.T(
                                         'כאשר משתמשים בתרחיש מוכן מראש התוכנית מקליטה תמיד, לכם נשאר לבחור את התדירות           ')],
                                     [S_gui.T('\n')],
                                     [S_gui.T(
                                         'שהוא מסוג אקסל ישן. מומלץ להשתמש באפשרות של המרה לאקסל בהגדרות ליצירת קובץ אקסל חדש' +
                                         ' CSV ' + 'כאשר מקליטים נוצר קובץ')],
                                     [S_gui.T(
                                         'יש לשים לב כי אם לא מסיימים את ההקלטה לפני סגירת התוכנית עלולות להיות בעיות! בפרט עם קבצי אקסל חדשים')],
                                     [S_gui.T('\n')],
                                     [S_gui.T(
                                         'כל קובץ הקלטה מכיל בתוכו עמודה לכל משתנה ועמודת זמן, אשר מתעדת את התאריך והזמן של דגימת הנתונים')],
                                     [S_gui.T('\n')],
                                     [S_gui.T(
                                         'ניתן לשנות קבצי אקסל ולהמשיך להקליט אליהם אך ההקלטה תימשך אחרי השורה הכתובה האחרונה')],
                                     [S_gui.T(
                                         'לכן ממולץ שלא לערוך קבצים לפני סיום הקלטת הנתונים ואם עורכים בכל זאת מומלץ לא לעשות שינויים בשורות ריקות')],
                                     [S_gui.Button('יציאה', k='-Exit-')]],
                             element_justification='Right', keep_on_top=True).read(close=True)
            elif selection == '-V_help-':
                S_gui.Window(title='עזרה עם משתנים',
                             layout=[[S_gui.T('?כיצד קוראים נתונים')],
                                     [S_gui.T(
                                         'במסך הראשי מופיעים המשתנים השונים')],
                                     [S_gui.T(
                                         'לכל משתנה יחידת המידה המתאימה לו, כך הטמפרטורה בצלזיוס, הלחויות באחוזים וכו')],
                                     [S_gui.T(
                                         'כאשר התכנית מקליטה היא שומרת את כל הנתונים הללו ללא יחידות המידה')],
                                     [S_gui.Button('יציאה', k='-Exit-')]],
                             element_justification='Right', keep_on_top=True).read(close=True)
            elif selection == '-C_help-':
                S_gui.Window(title='עזרה עם שליטה על החממה',
                             layout=[[S_gui.T('?כיצד שולטים בחממה')],
                                     [S_gui.T('ניתן לשנות את התאורה בחממה ולהפעיל את הברזים השונים')],
                                     [S_gui.T('ערכי התאורה הם בערכי אדום - ירוק - כחול')],
                                     [S_gui.T('כמו כן ניתן לשלוט בחממה באינטרוולים קבועים מראש בעזרת יצירת תרחיש')],
                                     [S_gui.Button('יציאה', k='-Exit-')]],
                             element_justification='Right', keep_on_top=True).read(close=True)
            elif selection == '-S_help-':
                S_gui.Window(title='עזרה עם תרחישים',
                             layout=[[S_gui.T('?כיצד יוצרים תרחישים')],
                                     [S_gui.T('ישנה תוכנית נוספת אשר נועדה לעזור לכם לייצור תרחישים')],
                                     [S_gui.T('תרחיש הוא למעשה סדר פעולות שבעצם מתפעלת את החממה באופן אוטומטי')],
                                     [S_gui.T('תרחישים הם למעשה קבצי טקסט שנוצרו על ידי התוכנית ליצירת תרחישים')],
                                     [S_gui.T(
                                         'תרחישים יכולים לשלוט בברזים, בתאורה ואפילו להשהות את ההתקדמות שלהם כדי שתוכלו לתפעל את החממה מבלי להיות לידה')],
                                     [S_gui.T('\n')],
                                     [S_gui.T('\n')],
                                     [S_gui.T('שימו לב! תרחישים מתחילים את עצמם מחדש אחרי 24 שעות, ללא קשר לאורכם',
                                              text_color='red')],
                                     [S_gui.T('\n')],
                                     [S_gui.T('\n')],
                                     [S_gui.T(
                                         'מבולבלים לגבי מה התרחיש שבחרתם עושה? מימין למעלה יופיעו באדום הפעולות של התרחיש')],
                                     [S_gui.Button('יציאה', k='-Exit-')]],
                             element_justification='Right', keep_on_top=True).read(close=True)
        help_window.close()

    window['-TIME-'].update(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    window['-PH-'].update(str(list_v[4]))
    window['-hum-'].update(str(list_v[1]) + '%')
    window['-gHum-'].update(str(list_v[3]) + '%')
    window['-light-'].update(str(list_v[2]) + ' LUX')
    window['-temp-'].update(str(list_v[0]) + 'C')

    input_s = input_e = input_f = ''
    # Handle serial communications
    while ser.inWaiting():
        input_s = ser.readline()
        input_e = input_s.decode('utf8', 'strict')
        input_f += input_e
        if input_f[0] == 'r':
            read = True

    if read:
        k = 0
        while k < len(input_f):
            if input_f[k] == ']':
                good = True
            k = k + 1
        if good:
            input_f = input_f[1:]
            input_f = input_f.replace(']', '\n')
            input_f += '\n'
            i = j = 0
            while i < 3:
                input_e = ''
                while j < len(input_f) and input_f[j] != ',' and input_f[j] != '\n':
                    input_e += input_f[j]
                    j = j + 1
                j = j + 1
                list_v[i] = float(input_e)
                i = i + 1
            good = False
        read = False

    if run_through_scenario:
        if delay_timer <= time.time():
            apply_scenario()

    if faucet1_timer_active:
        if faucet1_timer <= time.time():
            close_faucet1()
            faucet1_timer_active = False
    if faucet2_timer_active:
        if faucet2_timer <= time.time():
            close_faucet2()
            faucet2_timer_active = False

    if faucet1_cond_active:
        if mode_dic[faucet1_mode]:
            if list_v[cond_dic[faucet1_parameter]] < faucet1_stop_value:
                close_faucet1()
                faucet1_cond_active = False
        elif list_v[cond_dic[faucet1_parameter]] > faucet1_stop_value:
            close_faucet1()
            faucet1_cond_active = False

    if faucet2_cond_active:
        if mode_dic[faucet2_mode]:
            if list_v[cond_dic[faucet2_parameter]] < faucet2_stop_value:
                close_faucet2()
                faucet2_cond_active = False
        elif list_v[cond_dic[faucet2_parameter]] > faucet2_stop_value:
            close_faucet2()
            faucet2_cond_active = False

    if writing_entries:
        if recording_timer < time.time():
            recording_file.write(
                datetime.now().strftime("%d/%m/%Y %H:%M:%S") + ',' + str(list_v[0]) + ',' + str(list_v[1]) + ',' + str(
                    list_v[2]) + ',' + str(list_v[3]) + ',' + str(list_v[4]) + '\n')
            input_s = input_e = input_f = ""
            add_to_file = False
            recording_timer = time.time() + (recording_rate * 60)

window.close()
